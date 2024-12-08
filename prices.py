import openpyxl
from seleniumbase import Driver
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Load the Excel file
excel_path = "prices.xlsx"  # Path to your Excel file
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb.active

driver = Driver(uc=True)

url = "https://www.backmarket.es/es-es"
driver.uc_open_with_reconnect(url, 4)

accept_cookies = WebDriverWait(driver, 15).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="__nuxt"]/div/div[3]/div/div[2]/section/div/div/div[2]/button[3]/div'))
    )
accept_cookies.click()

for i in range(1, 4):
    row_index = i + 2
    search_term = (f"{sheet.cell(row=row_index, column=3).value} {sheet.cell(row=row_index, column=4).value}".strip()
                   .replace('BLACK', 'Negro')
                   .replace('BLUE', 'Azul')
                   .replace('BROWN', 'Marrón')
                   .replace('BURGUNDY', 'Burdeos')
                   .replace('CORAL', 'Coral')
                   .replace('CREAM', 'Crema')
                   .replace('GOLD', 'Oro')
                   .replace('GRAPHITE', 'Grafito')
                   .replace('GRAY', 'Gris')
                   .replace('GREEN', 'Verde')
                   .replace('LAVENDER', 'Lavanda')
                   .replace('LIME', 'Lima')
                   .replace('MINT', 'Menta')
                   .replace('NAVY', 'Azul marino')
                   .replace('OLIVE', 'Oliva')
                   .replace('ORANGE', 'Naranja')
                   .replace('PINK', 'Rosa')
                   .replace('PURPLE', 'Púrpura')
                   .replace('RED', 'Rojo')
                   .replace('SILVER', 'Plata')
                   .replace('SPACE GRAY', 'Gris espacial')
                   .replace('TITANIUM', 'Titanio')
                   .replace('VIOLET', 'Violeta')
                   .replace('WHITE', 'Blanco')
                   .replace('YELLOW', 'Amarillo'))

    # Find and clear the search box
    search_box = WebDriverWait(driver, 15).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="searchBar-input"]'))
    )
    search_box.send_keys(Keys.CONTROL + "a")
    search_box.send_keys(Keys.DELETE)
    search_box.send_keys(search_term, Keys.ENTER)

    # first search result
    result = WebDriverWait(driver, 15).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="__nuxt"]/div/div[4]/div[1]/div/div/section/div[2]/main/div[2]/a[1]/div/div[2]/div'))
    )
    result.click()

    # link
    link_element = WebDriverWait(driver, 15).until(
        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "a[data-qa='productCard']"))
    )
    link = link_element[0].get_attribute("href")

    WebDriverWait(driver, 15).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div/div[4]/div[4]/div/div/aside[1]/div/div[2]/div/ul/li[4]/button/div/div[2]/div/span[2]"))
    )
    premium = driver.find_element(By.XPATH, "/html/body/div[1]/div/div[4]/div[4]/div/div/aside[1]/div/div[2]/div/ul/li[4]/button/div/div[2]/div/span[2]").text.replace('€', '')
    excellent = driver.find_element(By.XPATH, "/html/body/div[1]/div/div[4]/div[4]/div/div/aside[1]/div/div[2]/div/ul/li[3]/button/div/div[2]/div/span[2]").text.replace('€', '')
    very_good = driver.find_element(By.XPATH, "/html/body/div[1]/div/div[4]/div[4]/div/div/aside[1]/div/div[2]/div/ul/li[2]/button/div/div[2]/div/span[2]").text.replace('€', '')
    correct = driver.find_element(By.XPATH, "/html/body/div[1]/div/div[4]/div[4]/div/div/aside[1]/div/div[2]/div/ul/li[1]/button/div/div[2]/div/span[2]").text.replace('€', '')

    # Save prices to Excel
    sheet.cell(row=row_index, column=5, value=premium)  # Column E
    sheet.cell(row=row_index, column=6, value=excellent)  # Column F
    sheet.cell(row=row_index, column=7, value=very_good)  # Column G
    sheet.cell(row=row_index, column=8, value=correct)  # Column H
    sheet.cell(row=row_index, column=9, value=link)  # Column I

    # Save Excel after every 10 entries (to avoid data loss in case of an interruption)
    if row_index % 10 == 0:
        wb.save(excel_path)

    print(f"Row {row_index} processed successfully.")

    # # Save changes to the Excel file
    # wb.save(excel_path)
    # wb.close()